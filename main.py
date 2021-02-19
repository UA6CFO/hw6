# Парсинг экселевского файла logs.xlsx и формирование файла отчёта report.xlsx
# Перед запуском удалите содержимое report.xlsx, но сам файл не удалять.
# --------------------------------
# Прошу ногами сильно не пинать за отсутствие цветной раскраски и отсутствия трендов !
# Если эо ОЧЕНЬ надо - могу заморочится, но на это ухлопаю много времени - интерес к учёбе пропадёт...
# --------------------------------
#
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from my_lib import my_func
import string


file_report = '.\\hw6\\report.xlsx'
file_logs = '.\\hw6\\logs.xlsx'
flag_find = False
rs = {1: '', 2: '', 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0}  # если писать как символ - excel выёживается...
itog = {1: '', 2: '', 3: '', 4: '', 5: '', 6: '', 7: '', 8: '', 9: '', 10: '', 11: '', 12: '', 13: '', 14: ''}  # это для формул, заголовков...
prod = {1: '', 2: '', 3: '', 4: '', 5: '', 6: '', 7: '', 8: '', 9: '', 10: '', 11: '', 12: '', 13: '', 14: '', 15: '', 16: ''}
prodc = {1: '', 2: '', 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0, 13: 0, 14: 0, 15: 0, 16: 0}
stroka = 'CDEFGHIJKLMN'
rs_copy = rs
prodc_copy = prodc
i = 0
sr = ''
el = ''


# Инициализируемся...
wb = load_workbook(filename = file_report, data_only = True)
sheet_report = wb['Лист1']
wbl = load_workbook(filename = file_logs)
sheet_log = wbl['log']

# делаем шапку для репорта 1-го блока
sheet_report.cell(column=1, row=2, value = 'Посетители веб-сайта')
sheet_report.merge_cells('C3:N3')
sheet_report['C3'] = 'количество посещений'
sheet_report.cell(column=1, row=4, value = 'Браузер')
sheet_report.cell(column=2, row=4, value = 'Тренд')
for i in range(0, 12):
    sheet_report.cell(column=i + 3, row=4, value = my_func.get_rus_str_moon(i+1))

# начинаем парсить logs.xlsx для 1-го блока
for str_log in range(2, sheet_log.max_row + 1):
    flag_find = False
    print('Итерация - ' + str(i))  # что бы не скучать...
    i = i + 1
    # здесь начинается заполнение report.xlsx
    dat = my_func.get_num_moon_long(str(sheet_log.cell(column=7, row=str_log).value))
    for row in range(5, sheet_report.max_row + 1):
        sr = ((sheet_report.cell(column=1, row=row).value).upper()).strip()
        el = ((sheet_log.cell(column=4, row=str_log).value).upper()).strip()
        if sr == el:
            # если такой браузер уже в наличии
            old_value = sheet_report.cell(column=dat + 2, row=row).value
            sheet_report.cell(column=dat + 2, row=row, value = int(old_value) + 1)
            old_value = ''
            flag_find = True
    if flag_find == False:
        # добавляем новый браузер
        rs_copy = rs
        rs_copy[1] = sheet_log.cell(column=4, row=str_log).value
        rs_copy[dat + 2] = 1
        sheet_report.append(rs_copy)
print('Строк обработано: ' + str(i)) # для пущего контроля 
# добавим 'ИТОГО:' для 1-го блока
start = 2 
stop = sheet_report.max_row
itog[2] = 'ИТОГО:'
for i in range(0, len(stroka)):
    sr = '=SUM({}:{})'.format(stroka[i] + str(start), stroka[i] + str(stop))
    itog[i+3] = sr
sheet_report.append(itog)

# начало 2-го блока
for i in range(1, 17):
    prod[i] = ''
for i in range(0, 3):
    sheet_report.append(prod)  # добавим три сторки для отступа
# делаем шапку для 2-го блока   
prod[1] = 'Популярные товары'
sheet_report.append(prod)
for i in range(1, 17):
    prod[i] = ''
sheet_report.append(prod)    
sr = 'C{}:N{}'.format(str(sheet_report.max_row), str(sheet_report.max_row))
el = 'C{}'.format(str(sheet_report.max_row))
sheet_report.merge_cells(sr)
sheet_report[el] = 'количество продаж'
prod[1] = 'Товар'
prod[2] = 'Тренд'
for i in range(0, 12):
    prod[i + 3] = my_func.get_rus_str_moon(i+1)
prod[15] = ''  # prod[15] = 'М'
prod[16] = ''  # prod[16] = 'Ж'
sheet_report.append(prod)
start = sheet_report.max_row  # номер строки заголовка

# начинаем парсить logs.xlsx для 2-го блока
i = 0
for str_log in range(2, sheet_log.max_row + 1):
    flag_find = False
    print('Итерация - ' + str(i))  # что бы не скучать...
    i = i + 1
    # здесь начинается заполнение report.xlsx - 2-й блок
    dat = my_func.get_num_moon_long(str(sheet_log.cell(column=7, row=str_log).value))
    gender = ((sheet_log.cell(column=2, row=str_log).value).upper()).strip()
    sr = sheet_log.cell(column=8, row=str_log).value
    spisok = sr.split(',')
    for element in spisok:
        for row in range(start, sheet_report.max_row + 1):
            sr = ((sheet_report.cell(column=1, row=row).value).upper()).strip()
            el = (element.upper()).strip()
            if sr == el:
                # если такой товар уже в наличии
                old_value = sheet_report.cell(column=dat + 2, row=row).value
                sheet_report.cell(column=dat + 2, row=row, value = int(old_value) + 1)
                old_value = ''
                if gender == 'Ж':
                    old_value = sheet_report.cell(column=16, row=row).value
                    sheet_report.cell(column=16, row=row, value = int(old_value) + 1)
                else:
                    old_value = sheet_report.cell(column=15, row=row).value
                    sheet_report.cell(column=15, row=row, value = int(old_value) + 1)
                old_value = ''  # лишний раз почистить - не помешает
                flag_find = True
        if flag_find == False:
            # добавляем новый товар
            prodc_copy = prodc  # подготовили для вставки данных
            prodc_copy[1] = element
            prodc_copy[dat + 2] = 1
            if gender == 'Ж':
                prodc_copy[16] = 1
            else:
                prodc_copy[15] = 1
            sheet_report.append(prodc_copy)
print('Строк обработано: ' + str(i)) # для пущего контроля 

# добавим 'ИТОГО:' для 2-го блока
stop = sheet_report.max_row  # конечная строка Продаж
itog[2] = 'ИТОГО:'
for i in range(0, len(stroka)):
    sr = '=SUM({}:{})'.format(stroka[i] + str(start), stroka[i] + str(stop))
    itog[i+3] = sr
sheet_report.append(itog)

# формируем "подвальную" статистику
for i in range(1, 17):
    prod[i] = ''
for i in range(0, 2):
    sheet_report.append(prod)  # добавим две сторки для отступа
prod[1] = 'Предпочтения'
sheet_report.append(prod)
prod[1] = ''
prod[2] = 'Наименование товара'
sheet_report.append(prod)
max_m = {1: 'Самый популярный товар среди мужчин', 2: '', 3: 0}
max_w = {1: 'Самый популярный товар среди женщин', 2: '', 3: 0}
min_m = {1: 'Самый невостребованный товар среди мужчин', 2: '', 3: 0}
min_w = {1: 'Самый невостребованный товар среди женщин', 2: '', 3: 0}
row = start + 1  # начальная инициализация, а то ноль наименьшее...
max_m[3] = sheet_report.cell(column=15, row=row).value  # men max
max_m[2] = sheet_report.cell(column=1, row=row).value  # men товар
min_m[3] = sheet_report.cell(column=15, row=row).value  # men min
min_m[2] = sheet_report.cell(column=1, row=row).value  # men товар
max_w[3] = sheet_report.cell(column=16, row=row).value  # women max
max_w[2] = sheet_report.cell(column=1, row=row).value  # women товар
min_w[3] = sheet_report.cell(column=16, row=row).value  # women min
min_w[2] = sheet_report.cell(column=1, row=row).value  # women товар
# ищем максимумы и минимумы
for row in range(start + 2, stop + 1):
    if int(max_m[3]) < sheet_report.cell(column=15, row=row).value:
        max_m[3] = sheet_report.cell(column=15, row=row).value  # men max
        max_m[2] = sheet_report.cell(column=1, row=row).value  # men товар
    if int(min_m[3]) >= sheet_report.cell(column=15, row=row).value:
        min_m[3] = sheet_report.cell(column=15, row=row).value  # men min
        min_m[2] = sheet_report.cell(column=1, row=row).value  # men товар
    if int(max_w[3]) < sheet_report.cell(column=16, row=row).value:
        max_w[3] = sheet_report.cell(column=16, row=row).value  # women max
        max_w[2] = sheet_report.cell(column=1, row=row).value  # women товар
    if int(min_w[3]) >= sheet_report.cell(column=16, row=row).value:
        min_w[3] = sheet_report.cell(column=16, row=row).value  # women min
        min_w[2] = sheet_report.cell(column=1, row=row).value  # women товар
# почистим "изнанку"    
max_m[3] = ''
max_w[3] = ''
min_m[3] = ''
min_w[3] = ''
# покладём и положим...
sheet_report.append(max_m)
sheet_report.append(max_w)
sheet_report.append(min_m)
sheet_report.append(min_w)
# почистим лишнее - мавр сделал своё дело...
for row in range(start + 1, stop + 1):
    sheet_report.cell(column=15, row=row, value = '')
    sheet_report.cell(column=16, row=row, value = '')
# теперь все вышеписанные упражнения (извращения) сохраним, для потомков наверное...
wb.save(filename = file_report)
print('--- THE END ---')
